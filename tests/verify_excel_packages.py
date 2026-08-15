import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


root = Path(sys.argv[1] if len(sys.argv) > 1 else "/tmp/weekflow-excel-verification")
files = sorted(root.glob("*.xlsx"))
assert len(files) == 7, files

report_files = {"05_Dashboard_Report_EN.xlsx", "06_Managed_Person_Report_EN.xlsx", "07_Report_To_Report_EN.xlsx"}

for file in files:
    assert zipfile.is_zipfile(file), file
    with zipfile.ZipFile(file) as archive:
        names = archive.namelist()
        lowered = "\n".join(names).lower()
        assert "vbaproject" not in lowered, file
        assert "externallinks" not in lowered, file
        assert "connections.xml" not in lowered, file
        assert not any(name.lower().endswith(".bin") for name in names), file
        content_types = archive.read("[Content_Types].xml").decode("utf-8")
        workbook_xml = archive.read("xl/workbook.xml").decode("utf-8")
        app_xml = archive.read("docProps/app.xml").decode("utf-8")
        ET.fromstring(content_types)
        ET.fromstring(workbook_xml)
        ET.fromstring(app_xml)
        assert "macroEnabled" not in content_types, file
        assert "codeName=" not in workbook_xml, file
        assert "<DocSecurity>0</DocSecurity>" in app_xml, file

    workbook = load_workbook(file, read_only=False, data_only=False, keep_links=False)
    assert workbook.sheetnames, file
    if file.name in report_files:
        assert workbook.sheetnames == ["Overall Dashboard", "Timeline Dashboard", "Progress History"], (file, workbook.sheetnames)
        for sheet in workbook.worksheets:
            assert sheet.freeze_panes is None, (file, sheet.title, sheet.freeze_panes)
    if file.name in {"01_Task_Import_Template_EN.xlsx", "02_Current_Task_Data_EN.xlsx"}:
        assert workbook.sheetnames == ["Task Import", "Progress History", "Instructions"], (file, workbook.sheetnames)
    workbook.close()

print("openpyxl/OOXML verification passed:", ", ".join(file.name for file in files))
