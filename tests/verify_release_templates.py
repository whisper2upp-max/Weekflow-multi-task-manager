import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TASK_TEMPLATE = ROOT / "templates" / "Weekflow_Task导入模板.xlsx"
DOCUMENT_TEMPLATE = ROOT / "templates" / "Weekflow_资料库导入模板.xlsx"


def verify_package(path, sheetnames):
    assert zipfile.is_zipfile(path), path
    with zipfile.ZipFile(path) as archive:
        names = archive.namelist()
        lowered = "\n".join(names).lower()
        assert "vbaproject" not in lowered, path
        assert "externallinks" not in lowered, path
        assert "connections.xml" not in lowered, path
        assert not any(name.lower().endswith(".bin") for name in names), path
        for name in ("[Content_Types].xml", "xl/workbook.xml", "docProps/app.xml"):
            ET.fromstring(archive.read(name))
        workbook_xml = archive.read("xl/workbook.xml").decode("utf-8")
        app_xml = archive.read("docProps/app.xml").decode("utf-8")
        assert "codeName=" not in workbook_xml, path
        assert "<DocSecurity>0</DocSecurity>" in app_xml, path
    workbook = load_workbook(path, read_only=False, data_only=False, keep_links=False)
    assert workbook.sheetnames == sheetnames, (path, workbook.sheetnames)
    assert workbook.properties.creator == "Wesley Yan", (path, workbook.properties.creator)
    workbook.close()


verify_package(TASK_TEMPLATE, ["Task导入", "进度历史", "填写说明"])
verify_package(DOCUMENT_TEMPLATE, ["资料库导入"])
print("Bundled v2.7 templates passed openpyxl/OOXML verification")
