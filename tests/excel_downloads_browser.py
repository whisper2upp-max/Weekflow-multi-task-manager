from pathlib import Path
import zipfile
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright


BASE_URL = "http://127.0.0.1:8765/Weekflow.html"
OUTPUT = Path("/tmp/weekflow-browser-downloads")
OUTPUT.mkdir(parents=True, exist_ok=True)

DATA = {
    "version": 3,
    "groups": [{"id": "g1", "name": "Service Delivery", "color": "#665CFF", "order": 1, "collapsed": False}],
    "flows": [{"id": "f1", "groupId": "g1", "name": "Review Flow", "color": "#665CFF", "order": 1, "collapsed": False}],
    "tasks": [{
        "id": "t1", "groupId": "g1", "flowId": "f1", "flowOrder": 1,
        "name": "Review Package", "reportTo": "Lucy Chen", "managedObject": "Jack Wang",
        "deliverable": "Approved package", "ddl": "2026-08-14", "urgency": "high",
        "status": "pending", "completedAt": None, "recurrenceCadence": "none",
        "recurrenceStart": None, "recurrenceEnd": None, "recurrenceCompletions": [],
        "progressNote": "Ready for review",
    }],
    "materials": [{
        "id": "m1", "title": "Review Guide", "url": "https://example.com/guide",
        "type": "document", "taskIds": ["t1"], "flowIds": ["f1"], "groupIds": ["g1"],
        "note": "Reference", "openEvents": [],
    }],
}


def download(page, trigger, expected_prefix):
    with page.expect_download() as info:
        trigger()
    item = info.value
    assert item.suggested_filename.startswith(expected_prefix), item.suggested_filename
    target = OUTPUT / item.suggested_filename
    item.save_as(target)
    return target


def verify_windows_safe(path, expected_sheets=None, require_unfrozen=False):
    assert zipfile.is_zipfile(path), path
    with zipfile.ZipFile(path) as archive:
        names = archive.namelist()
        lowered = "\n".join(names).lower()
        assert "vbaproject" not in lowered, path
        assert "externallinks" not in lowered, path
        assert "connections.xml" not in lowered, path
        assert not any(name.lower().endswith(".bin") for name in names), path
        content_types = archive.read("[Content_Types].xml").decode("utf-8")
        workbook_xml = archive.read("xl/workbook.xml").decode("utf-8")
        app_xml = archive.read("docProps/app.xml").decode("utf-8")
        ET.fromstring(content_types)
        ET.fromstring(workbook_xml)
        ET.fromstring(app_xml)
        assert "macroEnabled" not in content_types, path
        assert "codeName=" not in workbook_xml, path
        assert "<DocSecurity>0</DocSecurity>" in app_xml, path
    workbook = load_workbook(path, read_only=False, data_only=False, keep_links=False)
    if expected_sheets:
        assert workbook.sheetnames == expected_sheets, (path, workbook.sheetnames)
    if require_unfrozen:
        for sheet in workbook.worksheets:
            assert sheet.freeze_panes is None, (path, sheet.title, sheet.freeze_panes)
    workbook.close()


with sync_playwright() as playwright:
    browser = playwright.chromium.launch(headless=True)
    context = browser.new_context(viewport={"width": 1440, "height": 900}, accept_downloads=True)
    page = context.new_page()
    errors = []
    page.on("console", lambda message: errors.append(message.text) if message.type == "error" else None)
    page.on("pageerror", lambda error: errors.append(str(error)))
    page.goto(BASE_URL)
    page.evaluate("payload => localStorage.setItem('weekflow-v2.4:data:v4', JSON.stringify(payload))", DATA)
    page.evaluate("localStorage.setItem('weekflow-v2.4:language', 'en')")
    page.reload()
    page.wait_for_load_state("networkidle")

    page.get_by_role("button", name="Timeline", exact=True).click()
    page.wait_for_timeout(200)
    page.locator('.more-menu > summary').click()
    task_template = download(
        page,
        lambda: page.get_by_text("Download Excel Import Template", exact=True).click(),
        "Weekflow_Task_Import_Template_EN",
    )
    page.locator('.more-menu > summary').click()
    current_data = download(
        page,
        lambda: page.locator('[data-action="export-import-data"]').click(),
        "Weekflow_Current_Task_Data_",
    )
    dashboard = download(
        page,
        lambda: page.get_by_role("button", name="Export Dashboard Report").click(),
        "Task_Dashboard_",
    )

    page.get_by_role("button", name="Overall Dashboard", exact=True).click()
    page.get_by_role("button", name="Managed Person", exact=True).click()
    page.wait_for_timeout(200)
    managed = download(
        page,
        lambda: page.get_by_role("button", name='Export Task status for Managed Person “Jack Wang”').first.click(),
        "Managed_Person_Jack_Wang_Task_Status_",
    )
    page.get_by_role("button", name="Report To", exact=True).click()
    page.wait_for_timeout(200)
    report_to = download(
        page,
        lambda: page.get_by_role("button", name='Export Task status for Report To “Lucy Chen”').first.click(),
        "Report_To_Lucy_Chen_Task_Status_",
    )

    page.get_by_role("button", name="Document Library", exact=True).click()
    page.get_by_text("Download", exact=True).click()
    materials_template = download(
        page,
        lambda: page.locator('[data-template-kind="materials"]').click(),
        "Weekflow_Document_Import_Template_EN",
    )
    page.get_by_text("Download", exact=True).click()
    materials_library = download(
        page,
        lambda: page.locator('[data-action="export-materials"]').click(),
        "Weekflow_Document_Library_",
    )

    verify_windows_safe(task_template, ["Task Import", "Progress History", "Instructions"])
    verify_windows_safe(current_data, ["Task Import", "Progress History", "Instructions"])
    for report in (dashboard, managed, report_to):
        verify_windows_safe(
            report,
            ["Overall Dashboard", "Timeline Dashboard", "Progress History"],
            require_unfrozen=True,
        )
    verify_windows_safe(materials_template)
    verify_windows_safe(materials_library)

    assert not errors, errors
    browser.close()
    print("Browser downloads passed:", *(p.name for p in [task_template, current_data, dashboard, managed, report_to, materials_template, materials_library]))
